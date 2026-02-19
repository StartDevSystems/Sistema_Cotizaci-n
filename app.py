from flask import Flask, render_template, request, jsonify, make_response, Response
import firebase_admin
from firebase_admin import credentials, firestore
import datetime
from weasyprint import HTML
import os
from pathlib import Path
import io
import json
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Configuración de logs para Render
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# --- INICIALIZACIÓN DE FIREBASE ---
try:
    cred_file = "firebase_credentials.json"
    render_secret = "/etc/secrets/firebase_credentials.json"
    
    if os.path.exists(cred_file):
        cred = credentials.Certificate(cred_file)
        firebase_admin.initialize_app(cred)
        logger.info("Firebase: Cargado desde archivo local.")
    elif os.path.exists(render_secret):
        cred = credentials.Certificate(render_secret)
        firebase_admin.initialize_app(cred)
        logger.info("Firebase: Cargado desde /etc/secrets/.")
    elif os.environ.get("FIREBASE_JSON"):
        cred_dict = json.loads(os.environ.get("FIREBASE_JSON"))
        cred = credentials.Certificate(cred_dict)
        firebase_admin.initialize_app(cred)
        logger.info("Firebase: Cargado desde variable de entorno.")
    else:
        logger.error("Firebase: No se encontró configuración.")

    db = firestore.client()
except Exception as e:
    logger.error(f"Error fatal inicializando Firebase: {e}")

# -----------------------------------

def format_spanish_date(date_obj):
    months = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    return f"{date_obj.day} de {months[date_obj.month - 1]}, {date_obj.year}"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/clientes')
def clientes():
    return render_template('clientes.html')

@app.route('/cotizaciones')
def ver_cotizaciones():
    return render_template('cotizaciones.html')

# === RUTAS DE CLIENTES ===
@app.route('/api/clientes', methods=['GET'])
def listar_clientes():
    try:
        docs = db.collection('clientes').order_by('nombre').stream()
        clientes = []
        for doc in docs:
            data = doc.to_dict()
            clientes.append({
                'id': doc.id,
                'nombre': data.get('nombre', ''),
                'email': data.get('email', ''),
                'telefono': data.get('telefono', ''),
                'empresa': data.get('empresa', ''),
                'total_cotizaciones': data.get('total_cotizaciones', 0)
            })
        return jsonify({'status': 'ok', 'clientes': clientes})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/clientes', methods=['POST'])
def crear_cliente():
    try:
        data = request.get_json()
        cliente_data = {
            'nombre': data.get('nombre', ''),
            'email': data.get('email', ''),
            'telefono': data.get('telefono', ''),
            'empresa': data.get('empresa', ''),
            'total_cotizaciones': 0,
            'created_at': datetime.datetime.now(datetime.timezone.utc)
        }
        _, doc_ref = db.collection('clientes').add(cliente_data)
        return jsonify({'status': 'ok', 'cliente_id': doc_ref.id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# === RUTAS DE PRODUCTOS ===
@app.route('/api/productos', methods=['GET'])
def listar_productos():
    try:
        docs = db.collection('productos').order_by('nombre').stream()
        productos = [{'id': doc.id, **doc.to_dict()} for doc in docs]
        return jsonify({'status': 'ok', 'productos': productos})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/productos', methods=['POST'])
def crear_producto():
    try:
        data = request.get_json()
        producto_data = {
            'nombre': data.get('nombre', ''),
            'descripcion': data.get('descripcion', ''),
            'precio': float(data.get('precio', 0))
        }
        _, doc_ref = db.collection('productos').add(producto_data)
        return jsonify({'status': 'ok', 'producto_id': doc_ref.id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# === RUTAS DE COTIZACIONES ===
@app.route('/api/cotizaciones', methods=['GET'])
def listar_cotizaciones():
    try:
        docs = db.collection('cotizaciones').order_by('timestamp', direction=firestore.Query.DESCENDING).stream()
        cotizaciones = []
        for doc in docs:
            data = doc.to_dict()
            data['id'] = doc.id
            if data.get('timestamp'): data['timestamp'] = data['timestamp'].isoformat()
            cotizaciones.append(data)
        return jsonify({'status': 'ok', 'cotizaciones': cotizaciones})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/cotizaciones/<cotizacion_id>', methods=['GET'])
def obtener_cotizacion(cotizacion_id):
    try:
        doc = db.collection('cotizaciones').document(cotizacion_id).get()
        if not doc.exists: return jsonify({'status': 'error'}), 404
        data = doc.to_dict()
        data['id'] = doc.id
        return jsonify({'status': 'ok', 'cotizacion': data})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/guardar-cotizacion', methods=['POST'])
def guardar_cotizacion():
    try:
        data = request.get_json()
        counter_ref = db.collection('counters').document('cotizaciones_counter')
        
        @firestore.transactional
        def update_in_transaction(transaction, ref):
            snapshot = ref.get(transaction=transaction)
            new_number = snapshot.get('current_number') + 1
            transaction.update(ref, {'current_number': new_number})
            return new_number

        new_number = update_in_transaction(db.transaction(), counter_ref)
        data['quote_number'] = new_number
        data['timestamp'] = datetime.datetime.now(datetime.timezone.utc)
        data['estado'] = data.get('estado', 'Pendiente')
        _, doc_ref = db.collection('cotizaciones').add(data)

        cliente_nombre = data.get('cliente', '')
        if cliente_nombre:
            c_docs = db.collection('clientes').where('nombre', '==', cliente_nombre).limit(1).stream()
            for c in c_docs:
                db.collection('clientes').document(c.id).update({'total_cotizaciones': c.to_dict().get('total_cotizaciones', 0) + 1})

        return jsonify({'status': 'ok', 'message': f'Nº {new_number:03d} guardada', 'cotizacion_id': doc_ref.id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/cotizaciones/<cotizacion_id>', methods=['PUT'])
def actualizar_cotizacion(cotizacion_id):
    try:
        data = request.get_json()
        data['updated_at'] = datetime.datetime.now(datetime.timezone.utc)
        db.collection('cotizaciones').document(cotizacion_id).update(data)
        return jsonify({'status': 'ok', 'message': 'Actualizada'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/cotizaciones/<cotizacion_id>', methods=['DELETE'])
def eliminar_cotizacion(cotizacion_id):
    try:
        db.collection('cotizaciones').document(cotizacion_id).delete()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/cotizaciones/<cotizacion_id>/duplicar', methods=['POST'])
def duplicar_cotizacion(cotizacion_id):
    try:
        original = db.collection('cotizaciones').document(cotizacion_id).get().to_dict()
        counter_ref = db.collection('counters').document('cotizaciones_counter')
        
        @firestore.transactional
        def update_in_transaction(transaction, ref):
            snapshot = ref.get(transaction=transaction)
            new_number = snapshot.get('current_number') + 1
            transaction.update(ref, {'current_number': new_number})
            return new_number

        new_number = update_in_transaction(db.transaction(), counter_ref)
        nueva_data = {
            **original,
            'quote_number': new_number,
            'fecha': datetime.datetime.now().strftime('%Y-%m-%d'),
            'estado': 'Pendiente',
            'timestamp': datetime.datetime.now(datetime.timezone.utc),
            'notas': f"Duplicada de #{original.get('quote_number', 0):03d}"
        }
        _, new_doc = db.collection('cotizaciones').add(nueva_data)
        return jsonify({'status': 'ok', 'cotizacion_id': new_doc.id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# === ESTADÍSTICAS (COMPLETO PARA GRÁFICAS) ===
@app.route('/api/estadisticas', methods=['GET'])
def obtener_estadisticas():
    try:
        docs = list(db.collection('cotizaciones').stream())
        total = len(docs)
        monto_total = 0
        por_estado = {'Pendiente': 0, 'Aprobada': 0, 'Rechazada': 0, 'En Revisión': 0}
        por_mes = {}
        
        for doc in docs:
            data = doc.to_dict()
            # Calcular monto
            monto = float(data.get('total', '$0.00').replace('$', '').replace(',', ''))
            monto_total += monto
            # Por estado
            est = data.get('estado', 'Pendiente')
            por_estado[est] = por_estado.get(est, 0) + 1
            # Por mes
            fecha = data.get('fecha', '')
            if fecha:
                mes = fecha[:7] # YYYY-MM
                por_mes[mes] = por_mes.get(mes, 0) + 1
        
        return jsonify({
            'status': 'ok',
            'total_cotizaciones': total,
            'monto_total': monto_total,
            'promedio': monto_total / total if total > 0 else 0,
            'por_estado': por_estado,
            'por_mes': dict(sorted(por_mes.items())[-6:])
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# === GENERACIÓN DE DOCUMENTOS ===
@app.route('/descargar-pdf/<cotizacion_id>')
def descargar_pdf(cotizacion_id):
    try:
        doc = db.collection('cotizaciones').document(cotizacion_id).get()
        data = doc.to_dict()
        fecha_obj = datetime.datetime.strptime(data.get('fecha', ''), '%Y-%m-%d')
        data['fecha_formateada'] = format_spanish_date(fecha_obj)
        data['fecha_validez'] = format_spanish_date(fecha_obj + datetime.timedelta(days=15))
        
        # Ruta corregida para Docker y Render
        logo_path = os.path.join(os.getcwd(), 'static', 'img', 'logo.jpg')
        if os.path.exists(logo_path):
            data['logo_url'] = Path(logo_path).as_uri()
        else:
            logger.warning(f"No se encontró el logo en: {logo_path}")
            data['logo_url'] = ""
        
        pdf = HTML(string=render_template('cotizacion_pdf.html', **data)).write_pdf()
        res = make_response(pdf)
        res.headers['Content-Type'] = 'application/pdf'
        res.headers['Content-Disposition'] = f'inline; filename=cotizacion_{data.get("quote_number", 0):03d}.pdf'
        return res
    except Exception as e:
        return str(e), 500

@app.route('/descargar-excel/<cotizacion_id>')
def descargar_excel(cotizacion_id):
    try:
        data = db.collection('cotizaciones').document(cotizacion_id).get().to_dict()
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotización"
        
        # Diseño básico de Excel
        header_fill = PatternFill(start_color="0D2A42", end_color="0D2A42", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = f"COTIZACIÓN Nº {data.get('quote_number', 0):03d}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A3'] = f"Cliente: {data.get('cliente', '')}"
        ws['A4'] = f"Fecha: {data.get('fecha', '')}"
        
        headers = ['Descripción', 'Cantidad', 'Precio', 'Total']
        ws.append([])
        ws.append(headers)
        for cell in ws[6]:
            cell.fill = header_fill
            cell.font = header_font
            
        for item in data.get('items', []):
            cant = item.get('cantidad', 0)
            prec = item.get('precio', 0)
            ws.append([item.get('descripcion', ''), cant, prec, cant * prec])
            
        ws.append(['', '', 'TOTAL GENERAL:', data.get('total', '')])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment;filename=cotizacion.xlsx'})
    except Exception as e:
        return str(e), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
